from django import template
from .decorators import login_required, permission_required
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.urls import reverse
import csv
from django.contrib.auth.models import Permission
from django.core.serializers import serialize
from django.http import HttpResponse


from django.db.models import Q

from .models import Enseignant
from .forms import EnseignantForm

from .models import Corps
from .forms import CorpsForm

from .models import Mention
from .forms import MentionForm

from .models import Grade
from .forms import GradeForm

from .models import Departement
from .forms import DepartementForm

from .models import Domaine
from .forms import DomaineForm

from .models import Situation
from .forms import SituationForm

from .models import Employe
from .forms import EmployeForm

from .models import Diplome
from .forms import DiplomeForm

from .models import TypeConge
from .forms import TypeCongeForm

from .models import Conge
from .forms import CongeForm


from .models import Fonction
from .forms import FonctionForm


from .models import Contrat
from .forms import ContratForm

from django.shortcuts import render, redirect, get_object_or_404
from .models import Categorie
from .forms import CategorieForm  # Assurez-vous d'avoir créé un formulaire correspondant

from .models import Affectation
from .forms import AffectationForm

from django.contrib import messages

from import_export.formats import base_formats

from import_export import resources

import pandas as pd
from .forms import ImportExcelForm

from tablib import Dataset
from .resources import EmployeResource
import openpyxl


from datetime import date

import pandas as pd
from datetime import date

import openpyxl
from openpyxl.styles import Border, Side
from django.utils.dateparse import parse_date
from pandas.api.types import is_datetime64_any_dtype as is_datetime
from pandas.api.types import is_numeric_dtype
from openpyxl.styles import Border, Side, Alignment

from pandas._libs.tslibs.nattype import NaTType


from django.http import JsonResponse
from django.http import HttpResponseBadRequest
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side
from datetime import datetime, timedelta
from io import BytesIO
from django.core.exceptions import ValidationError
from django.utils import timezone



from django.db.models import Count, Avg
from .models import Employe, Enseignant, Conge

























@login_required(login_url="/login/")
def index(request):
    context = {'segment': 'index'}
    
    # Récupération des données
    nombre_enseignants = Enseignant.objects.count()
    nombre_employes = Employe.objects.count()
    nombre_affectations = Affectation.objects.count()
    nombre_conges = Conge.objects.count()
    
    
    
    
    context['nombre_employes'] = nombre_employes 
    context['nombre_enseignants'] = nombre_enseignants
    context['nombre_affectations'] = nombre_affectations
    context['nombre_conges'] = nombre_conges
   

    return render(request, 'home/index.html', context)



@login_required(login_url="/login/")
def pages(request):
    context = {}
    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:

        load_template = request.path.split('/')[-1]

        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))
        context['segment'] = load_template

        html_template = loader.get_template('home/' + load_template)
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:

        html_template = loader.get_template('home/page-404.html')
        return HttpResponse(html_template.render(context, request))

    except:
        html_template = loader.get_template('home/page-500.html')
        return HttpResponse(html_template.render(context, request))





@permission_required('apps.view_categorie', raise_exception=True)
@login_required(login_url="/login/")
def liste_categories(request):
    categories = Categorie.objects.all()
    return render(request, 'categorie/liste_categories.html', {'categories': categories})



# Vue pour ajouter une nouvelle catégorie
@permission_required('apps.add_categorie', raise_exception=True)
@login_required(login_url="/login/")
def ajouter_categorie(request):
    user = request.user
    has_permission = user.has_perm('apps.add_categorie')  # Vérification de la permission
    if not has_permission:
        messages.error(request, "Vous n'avez pas la permission d'ajouter une catégorie.")
        return redirect('liste_categories')
    if request.method == 'POST':
        form = CategorieForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ajout avec succès.')
            return redirect('liste_categories')
    else:
        form = CategorieForm()
    
    context = {'form': form, 'has_permission': has_permission, 'user': user}
    return render(request, 'categorie/ajouter_categorie.html', context)



@permission_required('apps.change_categorie', raise_exception=True)
@login_required(login_url="/login/")
def modifier_categorie(request, id_categorie):
    categorie = get_object_or_404(Categorie, pk=id_categorie)
    if request.method == 'POST':
        form = CategorieForm(request.POST, instance=categorie)
        if form.is_valid():
            form.save()
            messages.success(request, 'Modification avec succès.')
            return redirect('liste_categories')
    else:
        form = CategorieForm(instance=categorie)
    return render(request, 'categorie/modifier_categorie.html', {'form': form, 'categorie': categorie})

# Vue pour supprimer une catégorie existante


@permission_required('apps.delete_categorie', raise_exception=True)
@login_required(login_url="/login/")
def supprimer_categorie(request, id_categorie):
    categorie = get_object_or_404(Categorie, pk=id_categorie)
    if request.method == 'POST':
        categorie.delete()
        messages.success(request, ' supprission avec succès.')
        return redirect('liste_categories')
    return render(request, 'categorie/supprimer_categorie.html', {'categorie': categorie})






@permission_required('apps.view_diplome', raise_exception=True)
@login_required(login_url="/login/")
def liste_diplomes(request):
    diplomes = Diplome.objects.all()
    return render(request, 'diplome/liste_diplomes.html', {'diplomes': diplomes})

# Vue pour ajouter un nouveau diplôme
@permission_required('apps.add_diplome', raise_exception=True)
@login_required(login_url="/login/")
def ajouter_diplome(request):
    if request.method == 'POST':
        form = DiplomeForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Ajout avec succès.')
                return redirect('liste_diplomes')
            except Exception as e:
                messages.error(request, f'Erreur lors de l\'ajout : {e}')
        else:
            messages.error(request, 'Le formulaire est invalide. Veuillez corriger les erreurs.')
            return redirect('liste_diplomes')
    else:
        form = DiplomeForm()

    return render(request, 'diplome/ajouter_diplome.html', {'form': form})


# Vue pour modifier un diplôme existant
@permission_required('apps.change_diplome', raise_exception=True)
@login_required(login_url="/login/")
def modifier_diplome(request, id_diplome):
    diplome = get_object_or_404(Diplome, pk=id_diplome)
    if request.method == 'POST':
        form = DiplomeForm(request.POST, instance=diplome)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, ' avecModification succès.')
                return redirect('liste_diplomes')
            except Exception as e:
                messages.error(request, f'Erreur lors de la Modification : {e}')
        else:
            messages.error(request, 'Le formulaire est invalide. Veuillez corriger les erreurs.')
    else:
        form = DiplomeForm(instance=diplome)
    return render(request, 'diplome/modifier_diplome.html', {'form': form, 'diplome': diplome})



# Vue pour supprimer un diplôme existant
@permission_required('apps.delete_diplome', raise_exception=True)
@login_required(login_url="/login/")
def supprimer_diplome(request, id_diplome):
    diplome = get_object_or_404(Diplome, pk=id_diplome)
    if request.method == 'POST':
        diplome.delete()
        messages.success(request, 'Supression avec succès.')
        return redirect('liste_diplomes')
    return render(request, 'diplome/supprimer_diplome.html', {'diplome': diplome})





@permission_required('apps.view_contrat', raise_exception=True)
@login_required(login_url="/login/")
def liste_contrats(request):
    contrats = Contrat.objects.all()
    return render(request, 'contrat/liste_contrats.html', {'contrats': contrats})



# Vue pour ajouter un nouveau contrat
@permission_required('apps.add_contrat', raise_exception=True)
@login_required(login_url="/login/")
def ajouter_contrat(request):
    ajout_reussi = False

    if request.method == 'POST':
        form = ContratForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ajout avec succès.')
            ajout_reussi = True

    else:
        form = ContratForm()

    if ajout_reussi:
        # Rediriger vers la liste des contrats en cas de succès
        return redirect('liste_contrats')

    return render(request, 'contrat/ajouter_contrat.html', {'form': form, 'ajout_reussi': ajout_reussi})


# Vue pour modifier un contrat existant
@permission_required('apps.change_contrat', raise_exception=True)
@login_required(login_url="/login/")
def modifier_contrat(request, id_contrat):
    contrat = get_object_or_404(Contrat, pk=id_contrat)
    if request.method == 'POST':
        form = ContratForm(request.POST, instance=contrat)
        if form.is_valid():
            form.save()
            messages.success(request, 'Modification avec succès.')
            return redirect('liste_contrats')
    else:
        form = ContratForm(instance=contrat)
    return render(request, 'contrat/modifier_contrat.html', {'form': form, 'contrat': contrat})

# Vue pour supprimer un contrat existant

@permission_required('apps.delet_contrat', raise_exception=True)
@login_required(login_url="/login/")
def supprimer_contrat(request, id_contrat):
    contrat = get_object_or_404(Contrat, pk=id_contrat)
    if request.method == 'POST':
        contrat.delete()
        messages.success(request, 'Suppression avec succès.')
        return redirect('liste_contrats')
    return render(request, 'contrat/supprimer_contrat.html', {'contrat': contrat})






def date_not_in_past(value):
    if value < timezone.now().date():
        raise ValidationError('La date de début ne peut pas être dans le passé.')






# Vue pour afficher la liste des congés (Read)
@permission_required('apps.view_conge', raise_exception=True)
@login_required(login_url="/login/")
def liste_conges(request):
    conges = Conge.objects.all()
    nombre_conges = Conge.objects.count()
    
    return render(request, 'conge/liste_conges.html', {'conges': conges})



# Vue pour ajouter un nouveau congé (Create)


# importez les modèles et autres dépendances nécessaires
@permission_required('apps.add_conge', raise_exception=True)
@login_required(login_url="/login/")
def ajouter_conge(request):
    # Récupérez la liste des employés, enseignants et types de congés depuis votre modèle
    employes = Employe.objects.all()
    enseignants = Enseignant.objects.all()
    typeconges = TypeConge.objects.all()

    if request.method == 'POST':
        # Traitez le formulaire d'ajout de congé
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        type_conge_id = request.POST.get('typeconge')
        employe_id = request.POST.get('employe')
        enseignant_id = request.POST.get('enseignant')

        # Validez les données et effectuez la logique de gestion des congés ici
        # Vous devez ajouter des vérifications pour vous assurer que les données sont valides.
        # Par exemple, assurez-vous que les dates sont au bon format et que les ID existent.

        if date_debut and date_fin and type_conge_id and (employe_id or enseignant_id):
            # Enregistrez le congé dans la base de données
            Conge.objects.create(
                date_debut=date_debut,
                date_fin=date_fin,
                type_conge_id=type_conge_id,
                employe_id=employe_id,
                enseignant_id=enseignant_id
            )
            messages.success(request, 'Ajout avec succès.')
            # Redirigez l'utilisateur vers la liste des congés ou une autre page appropriée
            return redirect('liste_conges')
        else:
            messages.error(request, 'Veuillez remplir tous les champs obligatoires.')

    return render(request, 'conge/ajouter_conge.html', {
        'employes': employes,
        'enseignants': enseignants,
        'typeconges': typeconges
    })
      
@permission_required('apps.change_conge', raise_exception=True)
@login_required(login_url="/login/")
def modifier_conge(request, id_conge):
    conge = get_object_or_404(Conge, pk=id_conge)
    employes = Employe.objects.all()
    enseignants = Enseignant.objects.all()
    typeconges = TypeConge.objects.all()

    if request.method == 'POST':
        form = CongeForm(request.POST, instance=conge)

        if form.is_valid():
            # Assurez-vous d'associer les valeurs récupérées au formulaire ici
            form.save()
            messages.success(request, 'Modification effectuée avec succès.')
            return redirect('liste_conges')  # Rediriger vers la liste des congés après la modification

    else:
        form = CongeForm(instance=conge)

    return render(request, 'conge/modifier_conge.html', {
        'form': form,
        'conge': conge,
        'employes': employes,
        'enseignants': enseignants,
        'typeconges': typeconges
    })

# Vue pour supprimer un congé existant (Delete)
@permission_required('apps.delete_conge', raise_exception=True)
@login_required(login_url="/login/")
def supprimer_conge(request, id_conge):
    conge = get_object_or_404(Conge, pk=id_conge)
    if request.method == 'POST':
        conge.delete()
        messages.success(request, 'Suppression avec succès.')
        return redirect('liste_conges')
    return render(request, 'conge/supprimer_conge.html', {'conge': conge})

# Vue pour afficher la liste des types de congés (Read)
@permission_required('apps.view_typeconge', raise_exception=True)
@login_required(login_url="/login/")
def liste_typeconges(request):
    typeconges = TypeConge.objects.all()
    return render(request, 'typeconge/liste_typeconges.html', {'typeconges': typeconges})

# Vue pour ajouter un nouveau type de congé (Create)
@permission_required('apps.add_typeconge', raise_exception=True)
@login_required(login_url="/login/")
def ajouter_typeconge(request):
    if request.method == 'POST':
        form = TypeCongeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ajout avec succès.')
            return redirect('liste_typeconges')
    else:
        form = TypeCongeForm()
    return render(request, 'typeconge/ajouter_typeconge.html', {'form': form})

# Vue pour modifier un type de congé existant (Update)

@permission_required('apps.change_typeconge', raise_exception=True)
@login_required(login_url="/login/")
def modifier_typeconge(request, id_typeconge):
    typeconge = get_object_or_404(TypeConge, pk=id_typeconge)
    employes = Employe.objects.all()
    enseignants = Enseignant.objects.all()

    if request.method == 'POST':
        form = TypeCongeForm(request.POST, instance=typeconge)
        if form.is_valid():
            form.save()
            messages.success(request, 'Modification avec succès.')
            return redirect('liste_typeconges')
    else:
        form = TypeCongeForm(instance=typeconge)
    return render(request, 'typeconge/modifier_typeconge.html', {
        'form': form, 
        'typeconge': typeconge,
        'employes': employes,
        'enseignants': enseignants,

    })

# Vue pour supprimer un type de congé existant (Delete)
@permission_required('apps.delete_typeconge', raise_exception=True)
@login_required(login_url="/login/")
def supprimer_typeconge(request, id_typeconge):
    typeconge = get_object_or_404(TypeConge, pk=id_typeconge)
    if request.method == 'POST':
        typeconge.delete()
        messages.success(request, 'Suppression avec succès.')
        return redirect('liste_typeconges')
    return render(request, 'typeconge/supprimer_typeconge.html', {'typeconge': typeconge})




# Vue pour afficher la liste des grades
@permission_required('apps.view_grade', raise_exception=True)
@login_required(login_url="/login/")
def liste_grades(request):
    grades = Grade.objects.all()
    return render(request, 'grade/liste_grades.html', {'grades': grades})


# Vue pour ajouter un nouveau grade
@permission_required('apps.change_grade', raise_exception=True)
@login_required(login_url="/login/")
def ajouter_grade(request):
    if request.method == 'POST':
        form = GradeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ajout avec succès.')
            return redirect('liste_grades')
    else:
        form = GradeForm()
    return render(request, 'grade/ajouter_grade.html', {'form': form})

# Vue pour modifier un grade existant

@login_required(login_url="/login/")
def modifier_grade(request, id_grade):
    grade = get_object_or_404(Grade, pk=id_grade)
    if request.method == 'POST':
        form = GradeForm(request.POST, instance=grade)
        if form.is_valid():
            form.save()
            messages.success(request, 'Modification avec succès.')
            return redirect('liste_grades')
    else:
        form = GradeForm(instance=grade)
    return render(request, 'grade/modifier_grade.html', {'form': form, 'grade': grade})

# Vue pour supprimer un grade existant
@permission_required('apps.delete_grage', raise_exception=True)
@login_required(login_url="/login/")
def supprimer_grade(request, id_grade):
    grade = get_object_or_404(Grade, pk=id_grade)
    if request.method == 'POST':
        grade.delete()
        messages.success(request, 'Suppression avec succès.')
        return redirect('liste_grades')
    return render(request, 'grade/supprimer_grade.html', {'grade': grade})



@login_required(login_url="/login/")
def employe_profile(request, employe_id):
    employe = get_object_or_404(Employe, id=employe_id)
    image_url = employe.get_image_url()

    context = {
        'employe': employe,
        'image_url': image_url,
    }
    return render(request, 'employes/employe_profile.html', {'employe': employe})


@permission_required('apps.view_employe')
@login_required(login_url="/login/")
def liste_employes(request):
    employes = Employe.objects.all()
    nombre_employes = Employe.objects.count()
    
    
    return render(request, 'employes/liste_employes.html', {'employes': employes, 'nombre_employes': nombre_employes})

    

 # Assurez-vous que vous avez importé le formulaire

from django.contrib.auth.decorators import permission_required
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User

# Utilisation de la décoration @permission_required pour vérifier la permission avant l'accès à la vue
@permission_required('apps.add_employe')
@login_required(login_url="/login/")
def ajouter_employe(request):
    success_message = None
    form = EmployeForm()

    if request.method == 'POST':
        form = EmployeForm(request.POST, request.FILES)
        if form.is_valid():
            employe = form.save()
            print("Employe saved with ID:", employe.id)
            messages.success(request, 'Ajout avec succès.')
            return redirect('liste_employes')
        else:
            messages.error(request, 'Corrigez les champs.')
            print("Form is not valid. Errors:", form.errors)

    contrats = Contrat.objects.all()
    categories = Categorie.objects.all()
    diplomes = Diplome.objects.all()
    situations = Situation.objects.all()
    domaines = Domaine.objects.all()
    departements = Departement.objects.all()
    grades = Grade.objects.all()
    fonctions = Fonction.objects.all()

    # ... (autres données de contexte)

    return render(request, 'employes/ajouter_employe.html', {
        'form': form,
        'contrats': contrats,
        'categories': categories,
        'diplomes': diplomes,
        'situations': situations,
        'domaines': domaines,
        'departements': departements,
        'grades': grades,
        'fonctions': fonctions,
        'success_message': success_message  # Ajoutez la variable de succès à la contexte
    })




@permission_required('apps.change_employe')
@login_required(login_url="/login/")
def modifier_employe(request, id_em):
    employe = get_object_or_404(Employe, id=id_em)
    contrats = Contrat.objects.all()
    domaines = Domaine.objects.all()
    fonctions = Fonction.objects.all()
    diplomes = Diplome.objects.all()
    categories = Categorie.objects.all()
    departements = Departement.objects.all()
    situations = Situation.objects.all()
    grades = Grade.objects.all()
    
    if request.method == 'POST':
        form = EmployeForm(request.POST, request.FILES, instance=employe)
        if form.is_valid():
            form.save()
            messages.success(request, 'Modification avec succès.')
            return redirect('liste_employes')  # Redirigez vers la page de liste des employés après la modification
    else:
        form = EmployeForm(instance=employe)
        
    return render(request, 'employes/modifier_employe.html', {
        'form': form, 
        'employe': employe,
        'contrats': contrats, 
        'domaines': domaines,
        'fonctions': fonctions,
        'diplomes': diplomes,
        'categories': categories,
        'departements': departements,
        'situations': situations,
        'grades': grades,
    })





@permission_required('apps.delete_employe')
@login_required(login_url="/login/")
def supprimer_employe(request, id_em):
    employe = Employe.objects.get(id=id_em)
    employe.delete()
    messages.success(request, 'Suppression avec succès.')
    return redirect('liste_employes')









# Vue pour afficher la liste des situations
@permission_required('apps.view_situation', raise_exception=True)
@login_required(login_url="/login/")
def liste_situations(request):
    situations = Situation.objects.all()
    return render(request, 'situation/liste_situations.html', {'situations': situations})

# Vue pour ajouter une nouvelle situation
@permission_required('apps.cadd_situation', raise_exception=True)
@login_required(login_url="/login/")
def ajouter_situation(request):
    if request.method == 'POST':
        form = SituationForm(request.POST)
        if form.is_valid():
            # Vous n'avez pas besoin de vous soucier d'ajouter "MOIS" ici, car le formulaire le fait automatiquement
            form.save()
            messages.success(request, 'Ajout avec succès.')
            return redirect('liste_situations')
    else:
        form = SituationForm()

    return render(request, 'situation/ajouter_situation.html', {'form': form})

# Vue pour modifier une situation existante
@permission_required('apps.change_situation', raise_exception=True)
@login_required(login_url="/login/")
def modifier_situation(request, id_st):
    situation = get_object_or_404(Situation, pk=id_st)
    if request.method == 'POST':
        form = SituationForm(request.POST, instance=situation)
        if form.is_valid():
            form.save()
            messages.success(request, 'Modification avec succès.')
            return redirect('liste_situations')
    else:
        form = SituationForm(instance=situation)
    return render(request, 'situation/modifier_situation.html', {'form': form, 'situation': situation})

# Vue pour supprimer une situation existante
@permission_required('apps.delete_stuation', raise_exception=True)
@login_required(login_url="/login/")
def supprimer_situation(request, id_st):
    situation = get_object_or_404(Situation, pk=id_st)
    if request.method == 'POST':
        situation.delete()
        messages.success(request, 'Suppression avec succès.')
        return redirect('liste_situations')
    return render(request, 'situation/supprimer_situation.html', {'situation': situation})









#



@permission_required('apps.view_fonction', raise_exception=True)
@login_required(login_url="/login/")
def liste_fonctions(request):
    fonctions = Fonction.objects.all()
    return render(request, 'fonction/liste_fonctions.html', {'fonctions': fonctions})


@permission_required('apps.add_fonction', raise_exception=True)
@login_required(login_url="/login/")
def ajouter_fonction(request):
    if request.method == 'POST':
        form = FonctionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ajout avec succès.')
            return redirect('liste_fonctions')
    else:
        form = FonctionForm()
    return render(request, 'fonction/ajouter_fonction.html', {'form': form})

@permission_required('apps.change_fonction', raise_exception=True)
@login_required(login_url="/login/")
def modifier_fonction(request, id_fc):
    fonction = get_object_or_404(Fonction, pk=id_fc)
    if request.method == 'POST':
        form = FonctionForm(request.POST, instance=fonction)
        if form.is_valid():
            form.save()
            messages.success(request, 'Modification avec succès.')
            return redirect('liste_fonctions')
    else:
        form = FonctionForm(instance=fonction)
    return render(request, 'fonction/modifier_fonction.html', {'form': form, 'fonction': fonction})



@permission_required('apps.delete_fonction', raise_exception=True)
@login_required(login_url="/login/")
def supprimer_fonction(request, id_fc):
    fonction = get_object_or_404(Fonction, pk=id_fc)
    if request.method == 'POST':
        fonction.delete()
        messages.success(request, 'Suppression avec succès.')
        return redirect('liste_fonctions')
    return render(request, 'fonction/supprimer_fonction.html', {'fonction': fonction})





@permission_required('apps.view_mention', raise_exception=True)
@login_required(login_url="/login/")
def liste_mentions(request):
    mentions = Mention.objects.all()
    return render(request, 'mention/liste_mentions.html', {'mentions': mentions})

@permission_required('apps.add_mention', raise_exception=True)
@login_required(login_url="/login/")
def ajouter_mention(request):
    if request.method == 'POST':
        form = MentionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ajout avec succès.')
            return redirect('liste_mentions')
    else:
        form = MentionForm()
    return render(request, 'mention/liste_mentions.html', {'form': form})
    

@permission_required('apps.change_mention', raise_exception=True)
@login_required(login_url="/login/")
def modifier_mention(request, id_me):
    mention = get_object_or_404(Mention, pk=id_me)
    if request.method == 'POST':
        form = MentionForm(request.POST, instance=mention)
        if form.is_valid():
            form.save()
            messages.success(request, 'Modification avec succès.')
            return redirect('liste_mentions')
    else:
        form = MentionForm(instance=mention)
    return render(request, 'mention/modifier_mention.html', {'form': form, 'mention': mention})
   

@permission_required('apps.delete_mention', raise_exception=True)
@login_required(login_url="/login/")
def supprimer_mention(request, id_me):
    mention = get_object_or_404(Mention, pk=id_me)
    if request.method == 'POST':
        mention.delete()
        messages.success(request, 'Suppression avec succès.')
        return redirect('liste_mentions')
    return render(request, 'mention/supprimer_mention.html', {'mention': mention})
    







@permission_required('apps.view_corps', raise_exception=True)
@login_required(login_url="/login/")
def liste_corps(request):
    corps = Corps.objects.all()
    return render(request, 'corps/liste_corps.html', {'corps': corps})

@permission_required('apps.add_corps', raise_exception=True)
@login_required(login_url="/login/")
def ajouter_corps(request):
    if request.method == 'POST':
        form = CorpsForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ajout avec succès.')
            return redirect('liste_corps')
    else:
        form = CorpsForm()
    
    context = {'form': form}
    return render(request, 'corps/ajouter_corps.html', context)


@permission_required('apps.change_corps', raise_exception=True)
@login_required(login_url="/login/")
def modifier_corps(request, id_cr):
    corps = get_object_or_404(Corps, pk=id_cr)
    if request.method == 'POST':
        form = CorpsForm(request.POST, instance=corps)
        if form.is_valid():
            form.save()
            messages.success(request, 'Modification avec succès.')
            return redirect('liste_corps')
    else:
        form = CorpsForm(instance=corps)
    return render(request, 'corps/modifier_corps.html', {'form': form, 'corps': corps})

@permission_required('apps.delete_corps', raise_exception=True)
@login_required(login_url="/login/")
def supprimer_corps(request, id_cr):
    corps = get_object_or_404(Corps, pk=id_cr)
    if request.method == 'POST':
        corps.delete()
        messages.success(request, 'Suppression avec succès.')
        return redirect('liste_corps')
    return render(request, 'corps/supprimer_corps.html', {'corps': corps})



@permission_required('apps.view_enseignant', raise_exception=True)
@login_required(login_url="/login/")
def liste_enseignants(request):
    enseignants = Enseignant.objects.all()
    nombre_enseignants = Enseignant.objects.count()
  
    return render(request, 'enseignants/liste_enseignants.html', {'enseignants': enseignants, 'nombre_enseignants': nombre_enseignants })


@login_required(login_url="/login/")
def enseignant_profile(request, enseignant_id):
    enseignant = Enseignant.objects.get(pk=enseignant_id)
    image_url = enseignant.get_image_url()

    context = {
        'enseignant': enseignant,
        'image_url': image_url,
    }
    return render(request, 'enseignants/enseignant_profile.html', {'enseignant': enseignant})

@permission_required('apps.add_enseignant', raise_exception=True)
@login_required(login_url="/login/")
def ajouter_enseignant(request):
    success_message = None
    form = EnseignantForm()
    if request.method == 'POST':
        form = EnseignantForm(request.POST, request.FILES)
        if form.is_valid():
            enseignant = form.save()
            print("Enseignant saved with ID:", enseignant.id)
            messages.success(request, 'Ajout avec succès.')
            return redirect('liste_enseignants')
        else:
            print("Form is not valid. Errors:", form.errors)
    corps = Corps.objects.all()
    mentions = Mention.objects.all()
    diplomes = Diplome.objects.all()
    situations = Situation.objects.all()
    domaines = Domaine.objects.all()
    departements = Departement.objects.all()
    grades = Grade.objects.all()
    return render(request, 'enseignants/ajouter_enseignant.html', {
        'form': form,        
        'diplomes': diplomes,
        'situations': situations,
        'domaines': domaines,
        'departements': departements,
        'grades': grades,
        'mentions': mentions,
        'corps': corps,
        'success_message': success_message  # Ajoutez la variable de succès à la contexte
    })



@permission_required('apps.change_enseignant', raise_exception=True)
@login_required(login_url="/login/")
def modifier_enseignant(request, id_en):
    enseignant = get_object_or_404(Enseignant, id=id_en)
    corps = Corps.objects.all()
    mentions = Mention.objects.all()
    diplomes = Diplome.objects.all()
    situations = Situation.objects.all()
    domaines = Domaine.objects.all()
    departements = Departement.objects.all()
    grades = Grade.objects.all()
    success_message = None

    if request.method == 'POST':
        form = EnseignantForm(request.POST, request.FILES, instance=enseignant)
        if form.is_valid():
            form.save()
            messages.success(request, 'Modification avec succès.')
            return redirect('liste_enseignants')  # Redirection après modification
        else:
            messages.error(request, 'Il y a une erreur')
    else:
        form = EnseignantForm(instance=enseignant)
        
        
    return render(request, 'enseignants/modifier_enseignant.html', {
        'form': form,
        'enseignant': enseignant,
        'diplomes': diplomes,
        'situations': situations,
        'domaines': domaines,
        'departements': departements,
        'grades': grades,
        'mentions': mentions,
        'corps': corps,
        'success_message': success_message,
         
         })


@permission_required('apps.delete_enseignant', raise_exception=True)
@login_required(login_url="/login/")
def supprimer_enseignant(request, id_en):
    enseignant = Enseignant.objects.get(id=id_en)
    enseignant.delete()
    messages.success(request, 'Suppression avec succès.')
    return redirect('liste_enseignants')



# Vues pour le modèle Domaine

@permission_required('apps.view_domaine', raise_exception=True)
@login_required(login_url="/login/")
def liste_domaines(request):
    domaines = Domaine.objects.all()
    return render(request, 'domaine/liste_domaines.html', {'domaines': domaines})


@permission_required('apps.add_domaine', raise_exception=True)
@login_required(login_url="/login/")
def ajouter_domaine(request):
    if request.method == 'POST':
        form = DomaineForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ajout avec succès.')
            return redirect('liste_domaines')
    else:
        form = DomaineForm()
    return render(request, 'domaine/ajouter_domaine.html', {'form': form})


@permission_required('apps.change_domaine', raise_exception=True)
@login_required(login_url="/login/")
def modifier_domaine(request, id_dom):
    domaine = get_object_or_404(Domaine, id=id_dom)
    if request.method == 'POST':
        form = DomaineForm(request.POST, instance=domaine)
        if form.is_valid():
            form.save()
            messages.success(request, 'Modification avec succès.')
            return redirect('liste_domaines')
    else:
        form = DomaineForm(instance=domaine)
    return render(request, 'domaine/modifier_domaine.html', {'form': form, 'domaine': domaine})


@permission_required('apps.delete_domaine', raise_exception=True)
@login_required(login_url="/login/")
def supprimer_domaine(request, id_dom):
    domaine = get_object_or_404(Domaine, id=id_dom)
    domaine.delete()
    messages.success(request, 'Suppression avec succès.')
    return redirect('liste_domaines')


@permission_required('apps.view_departement', raise_exception=True)
@login_required(login_url="/login/")
def liste_departements(request):
    departements = Departement.objects.all()
    return render(request, 'departement/liste_departements.html', {'departements': departements})


@permission_required('apps.add_departement', raise_exception=True)
@login_required(login_url="/login/")
def ajouter_departement(request):
    if request.method == 'POST':
        form = DepartementForm(request.POST)
        if form.is_valid():
            # Enregistrement du département
            form.save()
            messages.success(request, 'Ajout avec succès.')
            return redirect('liste_departements')
    else:
        form = DepartementForm()
        

    return render(request, 'departement/ajouter_departement.html', {
        'form': form,
        
    })

@permission_required('apps.change_departement', raise_exception=True)
@login_required(login_url="/login/")
def modifier_departement(request, id_dep):
    departement = get_object_or_404(Departement, id=id_dep)
    if request.method == 'POST':
        form = DepartementForm(request.POST, instance=departement)
        if form.is_valid():
            form.save()
            messages.success(request, 'Modification avec succès.')
            return redirect('liste_departements')
    else:
        form = DepartementForm(instance=departement)
    return render(request, 'departement/modifier_departement.html', {'form': form, 'departement': departement})


@permission_required('apps.delete_departement', raise_exception=True)
@login_required(login_url="/login/")
def supprimer_departement(request, id_dep):
    departement = get_object_or_404(Departement, id=id_dep)
    departement.delete()
    messages.success(request, 'Suppression avec succès.')
    return redirect('liste_departements')




@permission_required('apps.view_affectation', raise_exception=True)
@login_required(login_url="/login/")
def list_affectations(request):
    affectations = Affectation.objects.all()
    nombre_affectations = Affectation.objects.count()
    return render(request, 'affectation/list_affectations.html', {'affectations': affectations})





@permission_required('apps.change_affectation', raise_exception=True)
@login_required(login_url="/login/")
def create_affectation(request):
    if request.method == 'POST':
        form = AffectationForm(request.POST)
        if form.is_valid():
            affectation_instance = form.save()  # Sauvegarde l'instance dans la base de données
            messages.success(request, 'Affectation ajoutée avec succès.')
            return redirect('list_affectations')
        else:
            messages.error(request, 'Il y a des erreurs dans le formulaire.')
    else:
        form = AffectationForm()

    employes = Employe.objects.all()
    domaines = Domaine.objects.all()
    departements = Departement.objects.all()
    fonctions = Fonction.objects.all()

    return render(request, 'affectation/create_affectation.html', {
        'form': form,
        'employes': employes,
        'domaines': domaines,
        'departements': departements,
        'fonctions': fonctions,
    })

@permission_required('apps.change_affectation', raise_exception=True)
@login_required(login_url="/login/")
def update_affectation(request, pk):
    affectation = get_object_or_404(Affectation, pk=pk)

    if request.method == 'POST':
        form = AffectationForm(request.POST, instance=affectation)
        if form.is_valid():
            affectation_instance = form.save()  # Sauvegarde l'instance dans la base de données
            messages.success(request, 'Affectation mise à jour avec succès.')

            # Mettre à jour les informations de l'employé lié à l'affectation modifiée
            employe = affectation_instance.employe
            if employe:
                employe.domaine = affectation_instance.n_domaine
                employe.departement = affectation_instance.n_departement
                employe.fonction = affectation_instance.n_fonction
                employe.save()

            return redirect('list_affectations')
        else:
            messages.error(request, 'Il y a des erreurs dans le formulaire.')
    else:
        form = AffectationForm(instance=affectation)

    employes = Employe.objects.all()
    domaines = Domaine.objects.all()
    departements = Departement.objects.all()
    fonctions = Fonction.objects.all()

    return render(request, 'affectation/update_affectation.html', {
        'form': form,
        'employes': employes,
        'domaines': domaines,
        'departements': departements,
        'fonctions': fonctions,
        'affectation': affectation,
    })

@permission_required('apps.delete_affectation', raise_exception=True)
@login_required(login_url="/login/")
def delete_affectation(request, pk):
    affectation = get_object_or_404(Affectation, pk=pk)
    
    if request.method == 'POST':
        affectation.delete()
        messages.success(request, 'Affectation supprimée avec succès.')
        return redirect('list_affectations')

    return render(request, 'affectation/list_affectations.html', {'affectation': affectation})








# views.py
def export_to_excel(request):
    employes = Employe.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employes"

    headers = ["Nom", "Prénom", "Sexe", "IM", "Date de Naissance", "Département", "Diplôme", "Catégorie",
               "Indice", "Fonction", "Date PS", "Statut", "Durée Contrat", "CIN", "Adresse Exacte",
               "Numéro de Téléphone"]

    # Style pour les en-têtes (majuscules et bordures)
    header_style = openpyxl.styles.NamedStyle(name="header_style")
    header_style.font = openpyxl.styles.Font(bold=True)
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    header_style.border = Border(
        left=Side(style="medium"), right=Side(style="medium"),
        top=Side(style="medium"), bottom=Side(style="medium")
    )

    # Appliquer le style aux en-têtes
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header.upper()
        cell.style = header_style

    # Style pour les cellules de données (bordures et centrage)
    data_style = openpyxl.styles.NamedStyle(name="data_style")
    data_style.alignment = Alignment(horizontal='center', vertical='center')
    data_style.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    row_num = 2
    for employe in employes:
        data_row = [employe.nom_em, employe.prenom_em, employe.get_sex_display(), employe.im, employe.date_nais.strftime('%Y-%m-%d'),
                    employe.departement.nom_dep, employe.diplome.type_dip, employe.categorie.num_cat,
                    employe.indice, employe.fonction.nom_fc, employe.date_PF.strftime('%Y-%m-%d'), employe.contrat.type_con,
                    employe.contrat.dure_con, employe.cin, employe.adresse, employe.num_tel]

        for col_num, cell_value in enumerate(data_row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.style = data_style

        row_num += 1

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Liste employes.xlsx"'
    wb.save(response)

    return response



def roman_to_int(roman):
    roman_numerals = {'I': 1, 'V': 5, 'X': 10, 'L': 50, 'C': 100, 'D': 500, 'M': 1000}
    result = 0
    i = 0
    while i < len(roman):
        if i + 1 < len(roman) and roman_numerals[roman[i]] < roman_numerals[roman[i + 1]]:
            result += roman_numerals[roman[i + 1]] - roman_numerals[roman[i]]
            i += 2
        else:
            result += roman_numerals[roman[i]]
            i += 1
    return result


def import_excel(request):
    if request.method == 'POST' and 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']

        if excel_file.name.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active

                for row in ws.iter_rows(min_row=2, values_only=True):
                    try:
                        # Conversion de la catégorie
                        categorie_value = row[7]  # Assuming it's in the 7th column (index 6)
                        try:
                            categorie_num = int(categorie_value)
                        except ValueError:
                            print(f"Erreur dans la conversion de la catégorie à la ligne {row}: Valeur incorrecte.")
                            continue

                        # Conversion de la date de naissance
                        date_naissance_value = row[4]  # Assuming it's in the 5th column (index 4)
                        try:
                            date_naissance = datetime.strptime(str(date_naissance_value), "%Y-%m-%d")
                        except ValueError:
                            print(f"Erreur dans la date de naissance à la ligne {row}: Format de date incorrect.")
                            continue

                        # Création de l'objet Employe avec les données extraites du fichier Excel
                        employe = Employe(
                            nom_em=str(row[1]),  # Assuming 1st column is for 'NOM'
                            prenom_em=str(row[2]),  # Assuming 2nd column is for 'PRÉNOMS'
                            sex=str(row[3]),  # Assuming 3rd column is for 'SEXE'
                            im=int(row[4]),  # Assuming 4th column is for 'IM'
                            date_nais=date_naissance,
                            adresse=str(row[15]),
                            cin=str(row[14]), 
                            date_PF=date_PF, # Assuming 6th column is for 'ADRESSE EXACTE'
                            num_tel=str(row[16]),  # Assuming 7th column is for 'NUMERO TELEPHONE'
                            indice=int(row[9]),  # Assuming 9th column is for 'INDICE'
                            categorie=Categorie.objects.get_or_create(num_cat=str(categorie_num))[0],
                            diplome=Diplome.objects.get_or_create(type_dip=str(row[7]))[0],  # Change the index as per your Excel column
                            contrat=Contrat.objects.get_or_create(type_con=str(row[12]), dure_con=str(row[13]))[0],  # Change the index as per your Excel column
                            situation=Situation.objects.get_or_create(type_st=str(row[10]))[0],  # Change the index as per your Excel column
                            domaine=Domaine.objects.get_or_create(nom_dom=str(row[11]))[0],  # Change the index as per your Excel column
                            departement=Departement.objects.get_or_create(nom_dep=str(row[6]))[0],  # Change the index as per your Excel column
                            grade=Grade.objects.get_or_create(num_classe=int(row[13][0]), num_echellon=int(row[13][-1]))[0],  # Change the index and slicing as per your Excel column
                            fonction=Fonction.objects.get_or_create(nom_fc=str(row[10]))[0]  # Change the index as per your Excel column
                        )

                        employe.save()

                    except Exception as e:
                        print(f"Une erreur s'est produite lors de l'importation de la ligne : {str(e)}")

                return redirect('liste_employes')

            except Exception as e:
                print(f"Une erreur s'est produite lors de l'importation : {str(e)}")
                return render(request, 'erreur.html', {'message': 'Une erreur s\'est produite lors de l\'importation.'})
        else:
            return render(request, 'erreur.html', {'message': 'Le fichier doit être au format .xlsx'})

    return render(request, 'employes/import_excel.html')









def excel_serial_date_to_string(serial_date):
    if isinstance(serial_date, int):
        # Correction pour les dates avant 1900 (Excel 1900 date windowing)
        if serial_date > 59:
            serial_date -= 1
        return (datetime(1899, 12, 30) + timedelta(days=serial_date)).date()
    return serial_date  # Si ce n'est pas un numéro de série Excel, retournez la valeur telle quelle

def export_enseignants(request):
    enseignants = Enseignant.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enseignants"

    headers = ["Nom", "Prénom", "Sexe", "IM", "Date de Naissance", "Département", "Diplôme", 
               "Indice", "Date PS", "Statut", "CIN", "Adresse", "Numéro de Téléphone"]

    # Style pour les en-têtes (majuscules et bordures)
    header_style = openpyxl.styles.NamedStyle(name="header_style")
    header_style.font = openpyxl.styles.Font(bold=True)
    header_style.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    header_style.border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style="medium"), right=openpyxl.styles.Side(style="medium"),
        top=openpyxl.styles.Side(style="medium"), bottom=openpyxl.styles.Side(style="medium")
    )

    # Appliquer le style aux en-têtes
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header.upper()
        cell.style = header_style

    # Style pour les cellules de données (bordures et centrage)
    data_style = openpyxl.styles.NamedStyle(name="data_style")
    data_style.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    data_style.border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style="thin"), right=openpyxl.styles.Side(style="thin"),
        top=openpyxl.styles.Side(style="thin"), bottom=openpyxl.styles.Side(style="thin")
    )

    row_num = 2
    for enseignant in enseignants:
        data_row = [enseignant.nom_en, enseignant.prenom_en, enseignant.get_sex_display(), enseignant.im,
                    enseignant.date_nais.strftime('%Y-%m-%d'), enseignant.departement.nom_dep, enseignant.diplome.type_dip,
                    enseignant.indice, enseignant.date_PF.strftime('%Y-%m-%d'),
                    enseignant.situation.type_st, enseignant.cin, enseignant.adress, enseignant.num_tel]

        for col_num, cell_value in enumerate(data_row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.style = data_style

        row_num += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Liste enseignants.xlsx"'
    return response







def import_enseignants(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        if excel_file.name.endswith('.xlsx'):
            df = pd.read_excel(excel_file)

            for index, row in df.iterrows():
                try:
                    # Créez un objet Enseignant avec les données de la ligne Excel
                    enseignant = Enseignant(
                        nom_en=row['Nom'],
                        prenom_en=row['Prénom'],
                        sex=row['Sexe'],
                        im=row['IM'],
                        date_nais=row['Date de Naissance'],
                        departement=Departement.objects.get_or_create(nom_dep=row['Département'])[0],
                        diplome=Diplome.objects.get_or_create(type_dip=row['Diplôme'])[0],
                        indice=row['Indice'],
                        date_PF=row['Date PS'],
                        situation=Situation.objects.get_or_create(type_st=row['Statut'])[0],
                        cin=row['CIN'],
                        adress=row['Adresse'],
                        num_tel=row['Numéro de Téléphone']
                    )

                    enseignant.save()
                except Exception as e:
                    print(f"Une erreur s'est produite lors de l'importation de la ligne {index}: {str(e)}")

            return redirect('liste_enseignants')
        else:
            return render(request, 'erreur.html', {'message': 'Le fichier doit être au format .xlsx'})

    return render(request, 'import_enseignants.html')











def get_sex_data(request):
    nombre_hommes = Employe.objects.filter(sex='M').count()
    nombre_femmes = Employe.objects.filter(sex='F').count()
    nombre_femes = Enseignant.objects.filter(sex='F').count()
    nombre_homes = Enseignant.objects.filter(sex='M').count()
    return JsonResponse({
        'hommes': nombre_hommes, 
        'femmes': nombre_femmes,
        'homes': nombre_homes,
        'femes': nombre_femes,
        })

def get_age_data(request):
    # Obtention des données d'âge pour le graphique
    age_data = Employe.objects.all().values_list('age', flat=True)
    age_dict = dict()
    for age in age_data:
        if age in age_dict:
            age_dict[age] += 1
        else:
            age_dict[age] = 1
    return JsonResponse(age_dict)

def get_dept_data(request):
    # Obtention des données par département pour le graphique
    dept_data = Employe.objects.values('departement__nom_dep').annotate(count=Count('id'))
    dept_dict = dict()
    for entry in dept_data:
        dept_dict[entry['departement__nom_dep']] = entry['count']
    return JsonResponse(dept_dict)

  



def get_employee_info(request, employee_id):
    employee = Employe.objects.get(pk=employee_id)
    data = {
        'domaine': employee.domaine.id,
        'departement': employee.departement.id,
        'fonction': employee.fonction.id if employee.fonction else ''  # Vérification facultative pour la fonction
    }
    return JsonResponse(data)

def get_employee_info(request, employe_id):
    employe = get_object_or_404(Employe, id=employe_id)
    data = {
        'grade': employe.grade,
        'contrat': employe.contrat,
        # Ajoutez d'autres champs si nécessaire
    }
    return JsonResponse(data)



def get_enseignant_info(request, enseignant_id):
    enseignant = get_object_or_404(Enseignant, id=enseignant_id)
    data = {
        'grade': serialize('json', [enseignant.grade])[1:-1],  # Utilisez serialize pour rendre le Grade sérialisable
        # Ajoutez d'autres champs si nécessaire
    }
    return HttpResponse(data, content_type='application/json')
















def export_affectations_to_excel(request):
    affectations = Affectation.objects.all()
  # Remplacez "VotreModeleDjango" par le nom correct de votre modèle d'affectations

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Affectations"

    headers = [ "EMPLOYE", "DATE D'AFFECTATION", "FONCTION", "DEPARTEMENT", "DOMAINE", "MOTIF"]

    # Style pour les en-têtes (majuscules, centrage et bordures)
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True)
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    header_style.border = Border(
        left=Side(style="medium"), right=Side(style="medium"),
        top=Side(style="medium"), bottom=Side(style="medium")
    )

    # Appliquer le style aux en-têtes
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header.upper()
        cell.style = header_style

    # Style pour les cellules de données (centrage et bordures)
    data_style = NamedStyle(name="data_style")
    data_style.alignment = Alignment(horizontal='center', vertical='center')
    data_style.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    row_num = 2
    for affectation in affectations:
        data_row = [
           
            f"{affectation.employe.nom_em} {affectation.employe.prenom_em}",
            affectation.date_aff,
            affectation.n_fonction.nom_fc,
            affectation.n_departement.nom_dep,
            affectation.n_domaine.nom_dom,
            affectation.motif
        ]

        for col_num, cell_value in enumerate(data_row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.style = data_style

        row_num += 1

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Liste affectations.xlsx"'
    wb.save(response)

    return response


    

########################################################################################

from django.contrib.auth import logout
from django.shortcuts import redirect
from django.views.decorators.http import require_POST

@require_POST
def custom_logout(request):
    logout(request)
    return redirect('logout')  # Remplacez 'home' par la route vers laquelle vous voulez rediriger après la déconnexion


##############################################################################################""

# from django.urls import path
# from . import views  # Assurez-vous que l'import est correct pour accéder à la vue

# urlpatterns = [
#     path('logout/', views.custom_logout, name='logout'),
#     # autres chemins
# ]

##############################################################################################

# <form action="{% url 'logout' %}" method="post">
#     {% csrf_token %}
#     <button type="submit">Déconnexion</button>
# </form>


#########################################################################################